import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

patH_1 = r'C:\my_github\mypytjon\example\獲利追蹤'


data_0 = ()
df_0 = pd.DataFrame(data_0)

data_1 = ([0, 0], [0.5, 0.2], [0.1, 0.2])
df_1 = pd.DataFrame(data_1, columns=['A', 'B'], index=['one', 'two', 'three'])

#xlsx檔:不存在時會錯誤
#Sheet:保留其他的Sheet，新增/覆蓋sheep1
#註解:if_sheet_exists="overlay"改成'error'就不會覆蓋
'''
patH_2 = 'test'
with pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', mode="a", engine="openpyxl",if_sheet_exists="error") as writer:
    df_1.to_excel(writer, sheet_name="Sheet1")
'''

#xlsx檔:不存在時會自動生成
#Sheet:刪除其他的Sheet，新增/覆蓋sheep1
#註解:if_sheet_exists="overlay"改成'error'就不會覆蓋
#註解:增加 mode="a" & if_sheet_exists="overlay"，可以保留其他檔案，但無法自動生成xlsx檔
'''
patH_2 = 'test'
file_path = patH_1 + '\\' + patH_2 + '.xlsx'
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df_0.to_excel(writer, sheet_name='Sheet1')#不存在時會自動生成
with pd.ExcelWriter(file_path,mode="a", engine='openpyxl',if_sheet_exists="overlay") as writer:
    df_1.to_excel(writer, sheet_name='Sheet1')
    writer.save()
    writer.close()
'''

#保留其他的Sheet，不覆蓋繼續加寫在同一個sheep，如果沒有檔案會幫忙新增，但沒有sheep無法執行
'''
patH_2 = 'test'
file_path = patH_1 + '\\' + patH_2 + '.xlsx'
writer = pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl')
try:
    book = openpyxl.load_workbook(file_path)
    writer.book = book
    sheet1 = writer.book['Sheet1']
    for row in dataframe_to_rows(df_1, index=True, header=True):
        sheet1.append(row)
except:
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_0.to_excel(writer, sheet_name='Sheet1')#不存在時會自動生成
        print('以自動生成')
writer.save()
writer.close()

'''


#保留其他的Sheet，不覆蓋繼續加寫在同一個sheep，如果沒有檔案會幫忙新增，但沒有sheep無法執行
'''
patH_2 = 'test_1'
Sheet = 'Sheet1'
file_path = patH_1 + '\\' + patH_2 + '.xlsx'

try:
    book = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    df_0.to_excel(writer, sheet_name=Sheet, index=True)
    writer.save()
    book = openpyxl.load_workbook(file_path)


# Append the second DataFrame to 'Sheet1' in the existing or newly created Excel file
writer = pd.ExcelWriter(file_path, engine='openpyxl')
writer.book = book
sheet1 = writer.book[Sheet]


# Write the second DataFrame to the existing worksheet starting from the determined row
for row in dataframe_to_rows(df_1, index=True, header=True):
    sheet1.append(row)

writer.save()
writer.close()
'''

#xlsx檔:不存在時會自動生成
#Sheet:保留其他的Sheet，加寫當前的sheep
#註解:以pd的方式儲存，只有在spyder能正常運作

patH_2 = 'test'
Sheet = 'Sheet1'
try:
    try:#如果有xlsx檔，也有Sheet
        book = openpyxl.load_workbook(patH_1 + '\\' + patH_2 + '.xlsx')
        writer = pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl',mode="a", if_sheet_exists="overlay")
        writer.book = book
        sheet1 = writer.book[Sheet]
        start_row = sheet1.max_row 
        df_1.to_excel(writer, sheet_name=Sheet,startrow=start_row,index=True, header=False)
        writer.save()
    except:#如果有xlsx檔，沒有Sheet
        with pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl',mode="a", if_sheet_exists="overlay") as writer:
            df_0.to_excel(writer, sheet_name=Sheet)#不存在時會自動生成
            df_1.to_excel(writer, sheet_name=Sheet,index=True, header=True)
            writer.save()
except:#如果沒有xlsx檔
    with pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl') as writer:
        df_0.to_excel(writer, sheet_name=Sheet)#不存在時會自動生成
        df_1.to_excel(writer, sheet_name=Sheet,index=True, header=True)
        writer.save()
writer.close()





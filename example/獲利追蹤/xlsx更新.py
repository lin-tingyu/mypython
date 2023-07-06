import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

patH_1 = r'C:\my_github\mypytjon\example\獲利追蹤'
patH_2 = 'test'


data_1 = ([9, 2], [5, 2], [10, 20])
df_1 = pd.DataFrame(data_1, columns=['A', 'B'], index=['one', 'two', 'three'])

#刪除其他的Sheet，新增/覆蓋sheep1(index不見了)
'''
file_path = patH_1 + '\\' + patH_2 + '.xlsx'
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df_1.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
'''
#刪除其他的Sheet，新增/覆蓋Sheet1(有index)
'''
writer = pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl')
df_1.to_excel(writer, sheet_name='Sheet1')
writer.save()
'''

#保留其他的Sheet，覆蓋/增加'Sheet1'(有index)
'''
book = openpyxl.load_workbook(patH_1 + '\\' + patH_2 + '.xlsx')
writer = pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl')
writer.book = book
df_1.to_excel(writer, sheet_name='Sheet2')
writer.save()
'''


#保留其他的Sheet，不覆蓋繼續加寫在同一個sheep，如果沒有檔案會幫忙新增，但沒有sheep無法執行
'''
file_path = patH_1 + '\\' + patH_2 + '.xlsx'

try:
    book = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    df_1.to_excel(writer, sheet_name='Sheet2', index=False)
    writer.save()
    book = openpyxl.load_workbook(file_path)


# Append the second DataFrame to 'Sheet1' in the existing or newly created Excel file
writer = pd.ExcelWriter(file_path, engine='openpyxl')
writer.book = book

# Get the 'Sheet1' worksheet
sheet1 = writer.book['Sheet2']

# Determine the starting row for appending the data
start_row = sheet1.max_row + 1

# Write the second DataFrame to the existing worksheet starting from the determined row
for row in dataframe_to_rows(df_1, index=False, header=False):
    sheet1.append(row)

writer.save()
'''





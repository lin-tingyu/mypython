import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

patH_1 = r'C:\my_github\mypytjon\example\獲利追蹤'
patH_2 = 'test_3'

book = openpyxl.load_workbook(patH_1 + '\\' + patH_2 + '.xlsx')
# Create the first DataFrame
data_0 = ([0.9, -0.2], [0.5, 0.2], [1, 2])
df_0 = pd.DataFrame(data_0, columns=['A', 'B'], index=['one', 'two', 'three'])


# Write the first DataFrame to Excel
writer = pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl')
df_0.to_excel(writer, sheet_name='Sheet1', index=False)
writer.save()

# Create the second DataFrame
data_1 = ([9, -2], [5, 2], [10, 20])
df_1 = pd.DataFrame(data_1, columns=['A', 'B'], index=['one', 'two', 'three'])

# Load the existing Excel file
book = openpyxl.load_workbook(patH_1 + '\\' + patH_2 + '.xlsx')

# Append the second DataFrame to 'Sheet1' in the existing Excel file
writer = pd.ExcelWriter(patH_1 + '\\' + patH_2 + '.xlsx', engine='openpyxl')
writer.book = book

# Get the 'Sheet1' worksheet
sheet1 = writer.book['Sheet1']

# Determine the starting row for appending the data
start_row = sheet1.max_row + 1

# Write the second DataFrame to the existing worksheet starting from the determined row
for row in dataframe_to_rows(df_1, index=False, header=False):
    sheet1.append(row)

writer.save()